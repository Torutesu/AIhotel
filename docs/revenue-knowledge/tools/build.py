# -*- coding: utf-8 -*-
"""
rules_data.py から以下を生成する:
  ../レビューシート.xlsx        配布用レビューシート（判定プルダウン＋コメント）
  ../レベニュー知識ベース.md    ルールベース形式の知識ベース（システム・AIコメント生成の参照元）
実行: python build.py   （要 openpyxl）
"""
import os
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

from rules_data import CATEGORIES, TERMS, VERSION, DATE, all_rules

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.dirname(HERE)
XLSX = os.path.join(OUT_DIR, "レビューシート.xlsx")
MD = os.path.join(OUT_DIR, "レベニュー知識ベース.md")

JP = "Yu Gothic"
INK = "26313C"
BRASS = "96742F"
LINE = "D9D4C9"
HEAD_FILL = PatternFill("solid", fgColor="EFECE4")
CAT_FILL = PatternFill("solid", fgColor="26313C")
YELLOW = PatternFill("solid", fgColor="FFF2B8")
NOTE_FILL = PatternFill("solid", fgColor="FFFBEB")
OK_FILL = PatternFill("solid", fgColor="E3EFE7")
FIX_FILL = PatternFill("solid", fgColor="F6E0D0")
KIND_FILL = {
    "定義": "E8EEF5", "計算式": "E5F0EC", "判定ルール": "FFF4DD", "アクション": "F3E9F7",
    "注意": "FBE9E7", "アンチパターン": "F1E4E4", "説明ルール": "E4F2F3",
}
thin = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def f(size=10, bold=False, color=INK, italic=False):
    return Font(name=JP, size=size, bold=bold, color=color, italic=italic)


def add_judge(ws, col_letter, first, last):
    dv = DataValidation(type="list", formula1='"問題なし,修正あり,わからない"', allow_blank=True)
    dv.error = "プルダウンから選択してください"
    ws.add_data_validation(dv)
    rng = f"{col_letter}{first}:{col_letter}{last}"
    dv.add(rng)
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"問題なし"'], fill=OK_FILL,
                                                  font=Font(name=JP, size=10, color="2E7D4F", bold=True)))
    ws.conditional_formatting.add(rng, CellIsRule(operator="equal", formula=['"修正あり"'], fill=FIX_FILL,
                                                  font=Font(name=JP, size=10, color="B3572A", bold=True)))


def build_xlsx(rules):
    wb = Workbook()

    # ---------------- 使い方
    ws = wb.active
    ws.title = "使い方"
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 3, "B": 26, "C": 62, "D": 12, "E": 40}.items():
        ws.column_dimensions[col].width = w
    ws["B2"] = "ホテルレベニューマネジメント知識ベース　レビューシート"
    ws["B2"].font = f(15, True)
    ws["B3"] = f"レビュー用ドラフト {VERSION}（{DATE}）｜ ルール {len(rules)} 項目 ＋ 用語 {len(TERMS)} 語"
    ws["B3"].font = f(10, color="6B7480")
    rows = [
        ("このシートについて",
         "AIレベニュー管理システムが需要予測・価格推奨・アラート・AIコメントの判断基準として参照する「レベニュー担当者の知識ベース」です。"
         "各行が1つのルール（定義・計算式・判定ルール・アクション等）になっています。実務経験に照らして、正しいか・現場感覚と合っているかをご確認ください。新しく書いていただくものはありません。"),
        ("記入方法",
         "①下の黄色セルにお名前を記入 → ②「ルール一覧」タブの各行を読み、判定列（黄色）でプルダウンから「問題なし / 修正あり / わからない」を選択 → "
         "③修正ありの場合はコメント列に修正案・理由（メモ書きでOK） → ④「用語集」「全体意見」タブも同様に。"),
        ("列の見方",
         "「種別」= そのルールの性質（定義／計算式／判定ルール／アクション／注意／アンチパターン／説明ルール）。"
         "「推奨パラメータ」= 汎用的な初期値の提案であり、実際のシステム設定値はホテルごとに調整します（この数値の妥当性も見ていただきたい点です）。"
         "「システム対応」= 本システムのどの機能・データに関係するか（参考情報。読み飛ばして構いません）。"),
        ("所要時間の目安", "約300行あります。60〜90分。優先順位があれば カテゴリ C（需要予測）・D（需要レベル）・E（価格決定）・Q（経験則）を先に。全部埋まらなくても構いません。"),
        ("返送方法", "記入後、このファイルをそのままメール等でご返送ください。複数名の回答は行のIDをキーに自動で集計します。"),
    ]
    r = 5
    for label, text in rows:
        ws.cell(row=r, column=2, value=label).font = f(10.5, True, BRASS)
        c = ws.cell(row=r, column=3, value=text)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
        c.font = f(10.5)
        c.alignment = WRAP
        ws.row_dimensions[r].height = 62
        r += 1
    r += 1
    ws.cell(row=r, column=2, value="お名前").font = f(10.5, True)
    c = ws.cell(row=r, column=3, value=""); c.fill = YELLOW; c.border = BORDER; c.font = f(10.5)
    r += 1
    ws.cell(row=r, column=2, value="記入日").font = f(10.5, True)
    c = ws.cell(row=r, column=3, value=""); c.fill = YELLOW; c.border = BORDER; c.font = f(10.5)
    c.number_format = "yyyy/mm/dd"
    r += 2
    ws.cell(row=r, column=2, value="記入例").font = f(10.5, True, BRASS)
    r += 1
    for i, h in enumerate(["項目", "内容", "判定", "コメント"]):
        c = ws.cell(row=r, column=2 + i, value=h)
        c.font = f(9.5, True); c.fill = HEAD_FILL; c.border = BORDER
    r += 1
    ex = ["残在庫による調整", "残在庫 ≤10% → +3ランク、≤20% → +2、≤30% → +1 …", "修正あり",
          "方向は正しいが、当館では残10%で+3は大きすぎる。+2が現実的。閑散期は残30%でも上げない方がよい"]
    for i, v in enumerate(ex):
        c = ws.cell(row=r, column=2 + i, value=v)
        c.font = f(9.5, italic=True, color="6B7480"); c.alignment = WRAP; c.border = BORDER
    ws.row_dimensions[r].height = 48
    r += 2
    ws.cell(row=r, column=2, value="回答状況").font = f(10.5, True, BRASS)
    r += 1
    ws.cell(row=r, column=2, value="ルール一覧 回答済み").font = f(10.5)
    ws.cell(row=r, column=3, value="=COUNTIF('ルール一覧'!G:G,\"問題なし\")+COUNTIF('ルール一覧'!G:G,\"修正あり\")+COUNTIF('ルール一覧'!G:G,\"わからない\")").font = f(10.5, True)
    ws.cell(row=r, column=4, value=f"/ {len(rules)}").font = f(10.5, color="6B7480")
    r += 1
    ws.cell(row=r, column=2, value="うち「修正あり」").font = f(10.5)
    ws.cell(row=r, column=3, value="=COUNTIF('ルール一覧'!G:G,\"修正あり\")").font = f(10.5, True, "B3572A")
    r += 2
    ws.cell(row=r, column=2, value="カテゴリ一覧").font = f(10.5, True, BRASS)
    r += 1
    cnt = Counter(x[1] for x in rules)
    for code, name, sysref in CATEGORIES:
        ws.cell(row=r, column=2, value=f"{code}　{name}").font = f(10)
        ws.cell(row=r, column=3, value=f"{cnt[code]} 項目 ｜ 関連: {sysref}").font = f(9.5, color="6B7480")
        r += 1

    # ---------------- ルール一覧
    ws2 = wb.create_sheet("ルール一覧")
    ws2.sheet_view.showGridLines = False
    for col, w in {"A": 7, "B": 16, "C": 22, "D": 11, "E": 74, "F": 26, "G": 12, "H": 44, "I": 30}.items():
        ws2.column_dimensions[col].width = w
    headers = ["ID", "カテゴリ", "項目", "種別", "内容（条件 → アクション → 根拠）", "推奨パラメータ（初期値）",
               "判定", "コメント（修正案・現場ではこうする 等）", "システム対応（参考）"]
    for i, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = f(10, True); c.fill = HEAD_FILL; c.border = BORDER; c.alignment = CENTER
    ws2.freeze_panes = "C2"
    ws2.row_dimensions[1].height = 32
    row = 2
    cur = None
    for rid, code, cat, item, kind, body, param, sysref in rules:
        if code != cur:
            cur = code
            sysref_cat = next(s for c_, n_, s in CATEGORIES if c_ == code)
            c = ws2.cell(row=row, column=1, value=f"{code}　{cat}　　（関連: {sysref_cat}）")
            ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
            c.font = Font(name=JP, size=10.5, bold=True, color="FFFFFF")
            c.fill = CAT_FILL
            c.alignment = Alignment(vertical="center")
            ws2.row_dimensions[row].height = 22
            row += 1
        vals = [rid, cat, item, kind, body, param, "", "", sysref]
        for i, v in enumerate(vals, start=1):
            cell = ws2.cell(row=row, column=i, value=v)
            cell.border = BORDER; cell.font = f(10); cell.alignment = WRAP
        ws2.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ws2.cell(row=row, column=1).font = f(9.5, True, BRASS)
        ws2.cell(row=row, column=2).font = f(9, color="6B7480")
        ws2.cell(row=row, column=3).font = f(10, True)
        kc = ws2.cell(row=row, column=4)
        kc.alignment = CENTER; kc.font = f(9)
        kc.fill = PatternFill("solid", fgColor=KIND_FILL.get(kind, "FFFFFF"))
        ws2.cell(row=row, column=6).font = f(9.5)
        ws2.cell(row=row, column=7).fill = YELLOW
        ws2.cell(row=row, column=7).alignment = CENTER
        ws2.cell(row=row, column=8).fill = NOTE_FILL
        ws2.cell(row=row, column=9).font = f(9, color="6B7480")
        lines = max(len(body) // 37 + 1, len(param) // 13 + 1, 2)
        ws2.row_dimensions[row].height = min(15 * lines + 6, 150)
        row += 1
    add_judge(ws2, "G", 2, row - 1)
    ws2.auto_filter.ref = f"A1:I{row - 1}"

    # ---------------- 用語集
    ws3 = wb.create_sheet("用語集")
    ws3.sheet_view.showGridLines = False
    for col, w in {"A": 6, "B": 32, "C": 60, "D": 12, "E": 40}.items():
        ws3.column_dimensions[col].width = w
    for i, h in enumerate(["No", "用語", "意味", "判定", "コメント（訂正・追加したい用語など）"], start=1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = f(10, True); c.fill = HEAD_FILL; c.border = BORDER; c.alignment = CENTER
    ws3.freeze_panes = "A2"
    row = 2
    for i, (term, mean) in enumerate(TERMS, start=1):
        for col, v in enumerate([i, term, mean, "", ""], start=1):
            cell = ws3.cell(row=row, column=col, value=v)
            cell.border = BORDER; cell.font = f(10); cell.alignment = WRAP
        ws3.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="top")
        ws3.cell(row=row, column=2).font = f(10, True)
        ws3.cell(row=row, column=4).fill = YELLOW
        ws3.cell(row=row, column=4).alignment = CENTER
        ws3.cell(row=row, column=5).fill = NOTE_FILL
        row += 1
    add_judge(ws3, "D", 2, row - 1)

    # ---------------- 全体意見
    ws4 = wb.create_sheet("全体意見")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 3
    ws4.column_dimensions["B"].width = 46
    ws4.column_dimensions["C"].width = 80
    ws4["B2"] = "全体を通してのご意見"
    ws4["B2"].font = f(13, True)
    qs = [
        "全体として実務感覚に合っていますか（合っている / 一部ズレあり など）",
        "カテゴリごと・テーマごと足りていないものがあれば",
        "推奨パラメータ（閾値・pt・ランク幅）で「これは違う」と感じた代表的なものと、ご自身の感覚値",
        "貴館の特殊事情（客層・立地・季節性）で一般ルールが当てはまらない点",
        "その他、自由記入",
    ]
    r = 4
    for q in qs:
        c = ws4.cell(row=r, column=2, value=q); c.font = f(10.5, True); c.alignment = WRAP
        a = ws4.cell(row=r, column=3, value=""); a.fill = YELLOW; a.border = BORDER; a.font = f(10.5); a.alignment = WRAP
        ws4.row_dimensions[r].height = 64
        r += 1

    wb.save(XLSX)


def build_md(rules):
    L = []
    L.append("# ホテルレベニューマネジメント知識ベース（ルールベース版）\n")
    L.append(f"**版**: {VERSION}（{DATE}）｜ ルール {len(rules)} 項目 ＋ 用語 {len(TERMS)} 語")
    L.append("**用途**: AIレベニュー管理システムの需要予測・需要レベル判定・価格推奨・アラート・AIコメント生成が参照する判断基準。"
             "また、汎用的な「レベニュー担当者の知識・判断」の原本として、説明・Q&Aの根拠に使う。")
    L.append("**生成元**: `tools/rules_data.py`（単一ソース）。`tools/build.py` で本MDと `レビューシート.xlsx` を生成する。手で本MDを編集せず、データ側を直す。")
    L.append("")
    L.append("> **実装値との関係**: 各ルールの「推奨パラメータ」は汎用的な初期値の提案。実際の閾値・重み・補正値はシステム設定（`PricingStrategyConfig`・料金ランク・ホテル設定）およびコード（`ruleBasedForecaster.ts` 等）が正であり、"
             "本書はそれらを**説明・検証・改善する根拠**として使う。両者が食い違う場合は実装を確認し、意図的な差ならデータ側の「システム対応」欄に注記する。")
    L.append("")
    L.append("## 目次\n")
    cnt = Counter(x[1] for x in rules)
    for code, name, sysref in CATEGORIES:
        L.append(f"- [{code}. {name}](#{code.lower()}-{name})（{cnt[code]}）— 関連: {sysref}")
    L.append("- [用語集](#用語集)")
    L.append("")
    L.append("## 種別の凡例\n")
    L.append("| 種別 | 意味 |\n|---|---|")
    L.append("| 定義 | 用語・指標・概念の定義 |")
    L.append("| 計算式 | 数式・算出手順 |")
    L.append("| 判定ルール | 条件 → 判断（閾値・分岐） |")
    L.append("| アクション | 条件が満たされたときに取る具体行動 |")
    L.append("| 注意 | データ・運用上の落とし穴 |")
    L.append("| アンチパターン | 典型的な失敗とその対策 |")
    L.append("| 説明ルール | AIコメント・画面表示で守る説明の型 |")
    L.append("")
    cur = None
    for rid, code, cat, item, kind, body, param, sysref in rules:
        if code != cur:
            cur = code
            sysref_cat = next(s for c_, n_, s in CATEGORIES if c_ == code)
            L.append(f"\n## {code}. {cat}\n")
            L.append(f"関連機能・データ: {sysref_cat}\n")
        L.append(f"### {rid} {item}")
        L.append(f"- **種別**: {kind}")
        L.append(f"- **内容**: {body}")
        if param and param != "—":
            L.append(f"- **推奨パラメータ（初期値）**: {param}")
        if sysref and sysref != "—":
            L.append(f"- **システム対応**: {sysref}")
        L.append("")
    L.append("\n## 用語集\n")
    L.append("| 用語 | 意味 |\n|---|---|")
    for term, mean in TERMS:
        L.append(f"| {term} | {mean} |")
    L.append("")
    with open(MD, "w", encoding="utf-8") as fp:
        fp.write("\n".join(L))


if __name__ == "__main__":
    rules = all_rules()
    build_xlsx(rules)
    build_md(rules)
    print(f"rules={len(rules)} terms={len(TERMS)}")
    print("wrote", XLSX)
    print("wrote", MD)
